#!/usr/bin/env python3
"""
Script to create a new Excel file with the proper structure.
This can be used to initialize or reset the Excel file.
"""

import os
import logging
import pandas as pd
from openpyxl import Workbook
from src.json_admin import load_json_config

# Configure logging
logging.basicConfig(level=logging.INFO, 
                   format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def create_excel_file(excel_path, sheet_mapping):
    """
    Create a new Excel file with the specified structure.
    
    Args:
        excel_path: Path where the Excel file should be created
        sheet_mapping: Dictionary with sheet configuration
        
    Returns:
        True if successful, False otherwise
    """
    try:
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        
        # Create a new workbook
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Create sheets based on sheet mapping
        for data_type, sheet_info in sheet_mapping.items():
            if isinstance(sheet_info, dict) and 'sheet_name' in sheet_info:
                sheet_name = sheet_info['sheet_name']
                # Create sheet
                ws = wb.create_sheet(sheet_name)
                
                # Add headers if column mapping exists
                if 'columns' in sheet_info and isinstance(sheet_info['columns'], dict):
                    headers = list(sheet_info['columns'].values())
                    if headers:
                        for col_idx, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col_idx, value=header)
        
        # Save the workbook
        wb.save(excel_path)
        logger.info(f"Created new Excel file: {excel_path}")
        return True
    except Exception as e:
        logger.error(f"Error creating Excel file: {str(e)}")
        return False

if __name__ == "__main__":
    # Get the base directory
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    # Load Excel configuration
    excel_config_path = os.path.join(base_dir, "config", "excel_sheet_mapping.json")
    excel_config = load_json_config(excel_config_path) or {}
    
    # Get Excel file path from config
    excel_path = excel_config.get("file_path", os.path.join(base_dir, "data", "report.xlsx"))
    
    # Get sheet mapping from config
    sheet_mapping = excel_config.get("sheets", {})
    
    # Check if file exists
    if os.path.exists(excel_path):
        logger.info(f"Excel file already exists at {excel_path}")
        backup_path = os.path.join(os.path.dirname(excel_path), "backups", 
                                  f"report_backup_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        os.makedirs(os.path.dirname(backup_path), exist_ok=True)
        
        try:
            # Create a backup
            import shutil
            shutil.copy2(excel_path, backup_path)
            logger.info(f"Created backup at {backup_path}")
            
            # Remove the existing file
            os.remove(excel_path)
            logger.info(f"Removed existing Excel file at {excel_path}")
        except Exception as e:
            logger.error(f"Error handling existing file: {str(e)}")
    
    # Create new Excel file
    success = create_excel_file(excel_path, sheet_mapping)
    
    if success:
        logger.info("Excel file created successfully")
    else:
        logger.error("Failed to create Excel file")
